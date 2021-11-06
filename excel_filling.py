import os
import openpyxl
import config
import game_scrapper as gs
import logging

def data_filling(appids, username):
    logger = logging.getLogger("logger")
    logger.debug(os.getcwd())
    workbook = openpyxl.load_workbook('data.xlsx')
    ws = workbook['Main']
    logger.info('Workbook opened')

    appids_existing = []
    for row in range(2, ws.max_row):
        appid = ws.cell(row=row, column=1).value
        if appid is None:
            continue
        appids_existing.append(int(appid))
    for appid in appids:
        logger.info('{} / {}'.format(appids.index(appid), len(appids)))
        if appid in appids_existing:
            logger.info('Skipped appid {}'.format(appid))
            username_cell = ws.cell(row=appids_existing.index(appid) + 1, column=config.EXCEL_MAPPER.get('username'))
            if username_cell.value is None or username in username_cell.value:
                continue
            username_cell.value += '{}, '.format(username)
        game = gs.get_game_info(appid)
        if game is None:
            continue
        gs.save_header(appid, game.get('name'))
        current_row = ws.max_row + 1
        for k, v in game.items():
            col_num = config.EXCEL_MAPPER.get(k)
            if col_num is None:
                continue
            # logger.info('Column - {}, value - {}'.format(k, v))
            ws.cell(row=current_row, column= col_num).value = v
        ws.cell(row=current_row, column= config.EXCEL_MAPPER.get('username')).value = username + ', '

    workbook.save('data.xlsx')
    workbook.close()
